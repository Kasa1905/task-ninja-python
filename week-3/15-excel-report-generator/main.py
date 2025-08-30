import pandas as pd
import argparse
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# --- CLI Arguments ---
def get_args():
    parser = argparse.ArgumentParser(description="Excel Report Generator: Reads raw data and generates a formatted Excel report.")
    parser.add_argument('--input', '-i', default='sample_sales_data.csv', help='Path to input CSV file')
    parser.add_argument('--output', '-o', default='sales_report.xlsx', help='Path to output Excel file')
    return parser.parse_args()

# --- Main Logic ---
def generate_report(input_csv, output_excel):
    # Read raw data
    df = pd.read_csv(input_csv)

    # Example summary: total sales by region
    summary = df.groupby('Region')['Sales'].sum().reset_index()
    summary.columns = ['Region', 'Total Sales']

    # Write to Excel with formatting
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Raw Data')
        summary.to_excel(writer, index=False, sheet_name='Summary')

        # Auto-format Summary sheet
        workbook = writer.book
        summary_sheet = writer.sheets['Summary']
        # Bold header, autofit columns, highlight top region
        for col_num, col_name in enumerate(summary.columns, 1):
            cell = summary_sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            column_width = max(12, len(col_name) + 2)
            summary_sheet.column_dimensions[get_column_letter(col_num)].width = column_width
        # Highlight top region
        if len(summary) > 0:
            max_row = summary['Total Sales'].idxmax() + 2
            for col_num in range(1, len(summary.columns)+1):
                summary_sheet.cell(row=max_row, column=col_num).fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

if __name__ == "__main__":
    args = get_args()
    generate_report(args.input, args.output)
    print(f"Report generated: {args.output}")
